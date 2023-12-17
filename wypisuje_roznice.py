import difflib
from openpyxl import load_workbook

sciezka_excela = 'test.xlsx'
nazwa_arkusza = "TC"

def znajdz_i_wpisz_roznice(tekst1, tekst2, sheet, row):
    differ = difflib.Differ()
    roznice = list(differ.compare(tekst1, tekst2))

    # Wyszukujemy zmienione znaki
    zmienione_znaki = [line[2:] for line in roznice if line.startswith('+ ')]
    zmienione_znaki = ' '.join(zmienione_znaki)

    # Wprowadzamy różnice do kolumny C
    sheet.cell(row=row, column=3, value=zmienione_znaki)

def zlicz_znaki(sheet, column):
    # Zliczamy ilość znaków w danej kolumnie
    tekst = [str(sheet.cell(row=row, column=column).value) for row in range(1, sheet.max_row + 1)]
    return sum(len(line) for line in tekst)

def oblicz_procent_niezmienionych(slowa_a, slowa_c):
    # Obliczamy procent niezmienionych słów w kolumnie B
    if slowa_a == 0:
        return 0
    return (slowa_a - slowa_c) / slowa_a * 100

def porownaj_i_wypisz_roznice(sciezka_excela, nazwa_arkusza):
    # Wczytujemy arkusz kalkulacyjny
    wb = load_workbook(sciezka_excela)

    # Wybieramy arkusz o danej nazwie
    sheet = wb[nazwa_arkusza]

    # Porównujemy dane z kolumn A i B i wprowadzamy różnice
    for row in range(1, sheet.max_row + 1):
        tekst1 = str(sheet.cell(row=row, column=1).value)
        tekst2 = str(sheet.cell(row=row, column=2).value)
        znajdz_i_wpisz_roznice(tekst1, tekst2, sheet, row)

    # Zliczamy ilość słów w kolumnie A po porównaniu
    slowa_a_po = zlicz_znaki(sheet, column=1)

    # Zliczamy ilość słów w kolumnie C (zmienione słowa)
    slowa_c = zlicz_znaki(sheet, column=3)

    # Zliczamy ilość różnic i wpisujemy do kolumny E w ostatnim wierszu
    sheet.cell(row=sheet.max_row+1, column=3, value=f'Ilość znaków z tekstu {slowa_a_po}')
    sheet.cell(row=sheet.max_row+1, column=3, value=f'Ilość znaków różnych {slowa_c}')

    # Obliczamy wskaźnik WRR i wpisujemy do kolumny F w ostatnim wierszu
    procent_niezmienionych = oblicz_procent_niezmienionych(slowa_a_po, slowa_c)
    sheet.cell(row=sheet.max_row+1, column=3, value=f"WRR: {procent_niezmienionych:.2f}%")

    # Zapisujemy zmiany do arkusza kalkulacyjnego
    wb.save(sciezka_excela)


# Porównujemy, wprowadzamy różnice i obliczamy ilość znaków oraz WRR
porownaj_i_wypisz_roznice(sciezka_excela, nazwa_arkusza)
